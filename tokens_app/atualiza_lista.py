import openpyxl as xl

def atualiza_lista_tokens():
    """
    'planilha': caminho do arquivo .xlsx (Digite o caminho sem o .xlsx)
    """
    tabela_novos_tokens = xl.load_workbook("tokens_app/ENTREGA_TOKENS_05.2025.xlsx")

    novos_tokens = {}

    for aba in tabela_novos_tokens.sheetnames:
        aba_atual = tabela_novos_tokens[aba]
        for linha in aba_atual.iter_rows(min_row=2, values_only=True):
            if linha[0] is not None:
                nome_responsavel = linha[0]
                cpf_responsavel = linha[1]
                funcao_responsavel = linha[2]
                serial = linha[3]
                data_solicitacao = linha[4]
                data_entrega = linha[5]
                observacao = linha[6]

                novos_tokens[f"{nome_responsavel}"] = {
                    'nome_responsavel': nome_responsavel,
                    'cpf_responsavel': cpf_responsavel,
                    'funcao_responsavel': funcao_responsavel,
                    'serial': serial,
                    'data_solicitacao': data_solicitacao,
                    'data_entrega': data_entrega,
                    'observacao': observacao
                }

    for token in novos_tokens:
        nome = novos_tokens[token]["nome_responsavel"]
        if nome[0] == " ":
            nome = nome[1:]
        if nome[-1] == " ":
            nome = nome[:-1]
        novos_tokens[token]["nome_responsavel"] = nome
        if novos_tokens[token]["observacao"] is None:
            novos_tokens[token]["observacao"] = ""

        # print(f"{novos_tokens[token]}\n")

    return novos_tokens

